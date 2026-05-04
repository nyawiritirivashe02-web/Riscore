"""
Export Utilities Module
Provides CSV and Excel export functionality for applications, anomalies, and reports.
"""

import csv
import io
import json
from datetime import datetime
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import logging

log = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to various formats."""

    @staticmethod
    def to_csv(data: List[Dict[str, Any]], filename: str = None) -> io.StringIO:
        """
        Export data to CSV format.
        
        Args:
            data: List of dictionaries to export
            filename: Optional filename (for reference only)
        
        Returns:
            StringIO object containing CSV data
        """
        if not data:
            return io.StringIO()
        
        try:
            output = io.StringIO()
            fieldnames = list(data[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            
            writer.writeheader()
            for row in data:
                writer.writerow(row)
            
            output.seek(0)
            return output
        except Exception as e:
            log.error(f"Error exporting to CSV: {e}")
            return io.StringIO()

    @staticmethod
    def to_excel(data: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> bytes:
        """
        Export data to Excel format.
        
        Args:
            data: List of dictionaries to export
            sheet_name: Name of the worksheet
        
        Returns:
            Bytes containing Excel file data
        """
        if not OPENPYXL_AVAILABLE:
            log.error("openpyxl not installed. Cannot export to Excel.")
            return b""
        
        if not data:
            return b""
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = row_data.get(header)
                    
                    # Format based on type
                    if isinstance(value, (int, float)):
                        cell.value = value
                        cell.alignment = Alignment(horizontal="right")
                    elif isinstance(value, datetime):
                        cell.value = value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        cell.value = str(value) if value is not None else ""
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            log.error(f"Error exporting to Excel: {e}")
            return b""

    @staticmethod
    def applications_to_csv(applications: List[Dict[str, Any]]) -> io.StringIO:
        """
        Export applications list to CSV with formatted columns.
        
        Args:
            applications: List of application dictionaries
        
        Returns:
            StringIO object containing CSV data
        """
        formatted_apps = []
        for app in applications:
            formatted_apps.append({
                "ID": app.get("id", ""),
                "Name": app.get("full_name", ""),
                "Loan Amount": app.get("loan_amount", 0),
                "Risk Score": app.get("risk_score", 0),
                "Risk Label": app.get("risk_label", ""),
                "Status": app.get("decision_status", ""),
                "Decision Date": app.get("decision_made_at", ""),
                "Credit Bureau": app.get("credit_bureau_covered", False),
                "Anomalies": app.get("anomaly_count", 0)
            })
        
        return ExportService.to_csv(formatted_apps, "applications.csv")

    @staticmethod
    def applications_to_excel(applications: List[Dict[str, Any]]) -> bytes:
        """
        Export applications list to Excel with formatting.
        
        Args:
            applications: List of application dictionaries
        
        Returns:
            Bytes containing Excel file data
        """
        formatted_apps = []
        for app in applications:
            formatted_apps.append({
                "ID": app.get("id", ""),
                "Name": app.get("full_name", ""),
                "Loan Amount": app.get("loan_amount", 0),
                "Risk Score": app.get("risk_score", 0),
                "Risk Label": app.get("risk_label", ""),
                "Status": app.get("decision_status", ""),
                "Decision Date": app.get("decision_made_at", ""),
                "Credit Bureau": app.get("credit_bureau_covered", False),
                "Anomalies": app.get("anomaly_count", 0)
            })
        
        return ExportService.to_excel(formatted_apps, "Applications")

    @staticmethod
    def anomalies_to_csv(anomalies: List[Dict[str, Any]]) -> io.StringIO:
        """
        Export anomalies to CSV with severity and timestamps.
        
        Args:
            anomalies: List of anomaly dictionaries
        
        Returns:
            StringIO object containing CSV data
        """
        formatted_anomalies = []
        for anomaly in anomalies:
            formatted_anomalies.append({
                "ID": anomaly.get("id", ""),
                "Borrower": anomaly.get("borrower_name", ""),
                "Type": anomaly.get("anomaly_type", ""),
                "Severity": anomaly.get("severity", ""),
                "Score": anomaly.get("anomaly_score", 0),
                "Description": anomaly.get("description", ""),
                "Detected": anomaly.get("detected_at", ""),
                "Resolved": anomaly.get("is_resolved", False)
            })
        
        return ExportService.to_csv(formatted_anomalies, "anomalies.csv")

    @staticmethod
    def anomalies_to_excel(anomalies: List[Dict[str, Any]]) -> bytes:
        """
        Export anomalies to Excel with formatting.
        
        Args:
            anomalies: List of anomaly dictionaries
        
        Returns:
            Bytes containing Excel file data
        """
        formatted_anomalies = []
        for anomaly in anomalies:
            formatted_anomalies.append({
                "ID": anomaly.get("id", ""),
                "Borrower": anomaly.get("borrower_name", ""),
                "Type": anomaly.get("anomaly_type", ""),
                "Severity": anomaly.get("severity", ""),
                "Score": anomaly.get("anomaly_score", 0),
                "Description": anomaly.get("description", ""),
                "Detected": anomaly.get("detected_at", ""),
                "Resolved": anomaly.get("is_resolved", False)
            })
        
        return ExportService.to_excel(formatted_anomalies, "Anomalies")

    @staticmethod
    def report_to_pdf_html(report_data: Dict[str, Any]) -> str:
        """
        Generate HTML for PDF report (can be converted to PDF using external library).
        
        Args:
            report_data: Dictionary containing report data
        
        Returns:
            HTML string
        """
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; border-bottom: 2px solid #3b82f6; padding-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th {{ background-color: #3b82f6; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                tr:hover {{ background-color: #f5f5f5; }}
                .summary {{ background-color: #f9fafb; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <h1>{report_data.get('title', 'Report')}</h1>
            <div class="summary">
                <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>{report_data.get('description', '')}</p>
            </div>
            {report_data.get('content_html', '')}
        </body>
        </html>
        """
        return html
